import pandas as pd
from openpyxl import load_workbook
import string
import xlsxwriter
import numpy

df = pd.read_excel('Distrib.xlsx')

#print(df[['CODIGO','cum']])
wb = load_workbook('Distrib.xlsx')
pestaña = wb['Hoja1']
min_colum = wb.active.min_column
max_colum = wb.active.max_column
min_fila = wb.active.min_row
max_fila = wb.active.max_row

#print(min_colum)
#print(max_colum)
#print(min_fila)
#print(max_fila)
df = df.loc[:,df.columns != 'PRODUCTO']

#df1 = df.iloc[:]
df1 =list(df.columns.values)
tienda = df1


cuenta_col = range(max_colum - 2)
print(tienda)
#tienda =["cum","qui","cal","dom","esm","ala","ova","lpa","sam","eco","sub","ira","cer","mai","vic","mac","pto","bel","rej"]
print(tienda,type(tienda))
for i in tienda:
    print(i)
    print(df[["CODIGO",i]])
    df2 = df[["CODIGO",i]]
    """workbook = xlsxwriter.Workbook("oc_"+i+".xlsx")
    worksheet = workbook.add_worksheet
    worksheet.write(df[["CODIGO",i]])
    writer = pd.ExcelWriter("G:/Mi unidad/Naturland-Monitor/PythonProjects/OC/oc_"+i+".xlsx")
    workbook.close()"""




